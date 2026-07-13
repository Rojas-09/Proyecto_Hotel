"""
Reporte Service - Lógica de negocio para reportes (RF-08)
Genera reportes de ocupación, ingresos y estadísticas en formato xlsx y pdf.
"""

import os
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, Image
from sqlalchemy import func, select

from app import db
from app.models.habitacion import Habitacion, TipoHabitacion
from app.models.huesped import Huesped
from app.models.pago import EstadoPago, Pago
from app.models.reserva import EstadoReserva, Reserva
# Matplotlib para gráficos en PDF
import matplotlib
matplotlib.use("Agg")  # Backend sin GUI


# ---------------------------------------------------------------------------
# Chart generation (matplotlib -> bytes)
# ---------------------------------------------------------------------------


def _grafico_ocupacion_por_tipo(labels, ocupaciones, ingresos) -> bytes:
    """Bar chart grouped: ocupación% + ingresos por tipo. Retorna PNG bytes."""
    import matplotlib.pyplot as plt

    fig, ax1 = plt.subplots(figsize=(8, 4))
    x = range(len(labels))
    width = 0.35

    ax1.bar(
        [i - width / 2 for i in x], ocupaciones, width,
        label="Ocupación %", color="#3498db",
    )
    ax1.set_ylabel("Ocupación %", color="#3498db")
    ax1.tick_params(axis="y", labelcolor="#3498db")
    max_ocup = max(ocupaciones) if ocupaciones else 0
    ax1.set_ylim(0, max(max_ocup * 1.2, 10))

    ax2 = ax1.twinx()
    ax2.bar(
        [i + width / 2 for i in x], ingresos, width,
        label="Ingresos ($)", color="#f39c12",
    )
    ax2.set_ylabel("Ingresos (COP)", color="#f39c12")
    ax2.tick_params(axis="y", labelcolor="#f39c12")

    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, rotation=45, ha="right")
    ax1.set_title("Ocupación e Ingresos por Tipo de Habitación")

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _grafico_tendencia_diaria(fechas, ocupaciones, adrs) -> bytes:
    """Line chart dual axis: ocupación% + ADR por día."""
    import matplotlib.dates as mdates
    import matplotlib.pyplot as plt

    fig, ax1 = plt.subplots(figsize=(8, 4))

    ax1.plot(fechas, ocupaciones, "o-", color="#3498db", label="Ocupación %")
    ax1.set_ylabel("Ocupación %", color="#3498db")
    ax1.tick_params(axis="y", labelcolor="#3498db")
    ax1.set_ylim(0, max(ocupaciones) * 1.2 if ocupaciones else 100)

    ax2 = ax1.twinx()
    ax2.plot(fechas, adrs, "s-", color="#f39c12", label="ADR (COP)")
    ax2.set_ylabel("ADR (COP)", color="#f39c12")
    ax2.tick_params(axis="y", labelcolor="#f39c12")

    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
    ax1.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(fechas) // 10)))
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha="right")

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")

    ax1.set_title("Tendencia Diaria: Ocupación y ADR")
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _grafico_top_habitaciones(labels, ingresos, noches) -> bytes:
    """Horizontal bar chart: Top 5 habitaciones por ingresos y noches."""
    import matplotlib.pyplot as plt

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4), sharey=True)

    y_pos = range(len(labels))

    ax1.barh(y_pos, ingresos, color="#27ae60")
    ax1.set_xlabel("Ingresos (COP)")
    ax1.set_title("Top 5 por Ingresos")
    ax1.invert_yaxis()

    ax2.barh(y_pos, noches, color="#2980b9")
    ax2.set_xlabel("Noches Ocupadas")
    ax2.set_title("Top 5 por Noches")
    ax2.invert_yaxis()

    for ax in (ax1, ax2):
        ax.set_yticks(y_pos)
        ax.set_yticklabels(labels, fontsize=9)

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _crear_directorio_reportes() -> str:
    """Crea y retorna la ruta del directorio de reportes."""
    directorio = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
        "static",
        "reportes",
    )
    if not os.path.exists(directorio):
        os.makedirs(directorio)
    return directorio


def _parse_fecha(fecha_str: str) -> date:
    """Convierte string YYYY-MM-DD a objeto date."""
    return datetime.strptime(fecha_str, "%Y-%m-%d").date()


def _fecha_str(fecha: date) -> str:
    """Convierte date a string YYYY-MM-DD."""
    return fecha.strftime("%Y-%m-%d")


def _formatear_numero(valor) -> float:
    """Convierte Decimal o número a float."""
    if isinstance(valor, Decimal):
        return float(valor)
    return float(valor)


def _fecha_default_mes_anterior():
    """Retorna (fecha_inicio, fecha_fin) del mes anterior completo."""
    hoy = date.today()
    primer_dia_mes_actual = hoy.replace(day=1)
    fin = primer_dia_mes_actual - timedelta(days=1)
    inicio = fin.replace(day=1)
    return _fecha_str(inicio), _fecha_str(fin)


# ---------------------------------------------------------------------------
# Helpers de generación de archivos
# ---------------------------------------------------------------------------


def _generar_xlsx(datos: list, nombre_archivo: str, encabezados: list) -> str:
    """
    Genera un archivo xlsx con datos y encabezados dados.
    Retorna la ruta del archivo creado.
    """
    directorio = _crear_directorio_reportes()
    ruta = os.path.join(directorio, nombre_archivo)

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_archivo.replace(".xlsx", "")[:31]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2C3E50", end_color="2C3E50", fill_type="solid"
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, encabezado in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_idx, value=encabezado)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    cell_align = Alignment(horizontal="left", vertical="center")
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = cell_align
            cell.border = border

    for col_idx in range(1, len(encabezados) + 1):
        max_length = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(datos) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

    wb.save(ruta)
    return ruta


def _generar_pdf(datos: list, nombre_archivo: str, titulo: str) -> str:
    """
    Genera un archivo pdf con una tabla de datos.
    Retorna la ruta del archivo creado.
    """
    directorio = _crear_directorio_reportes()
    ruta = os.path.join(directorio, nombre_archivo)

    doc = SimpleDocTemplate(
        ruta,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="TitleCenter",
            parent=styles["Heading1"],
            alignment=1,
            fontSize=16,
            spaceAfter=6,
        )
    )

    elements = []
    elements.append(Paragraph(titulo, styles["TitleCenter"]))
    elements.append(Spacer(1, 4 * mm))

    if not datos:
        elements.append(
            Paragraph("No hay datos disponibles para este reporte.", styles["Normal"])
        )
    else:
        encabezados = datos[0]
        filas = datos[1:]

        table_data = [encabezados] + filas
        col_count = len(encabezados)
        col_width = (210 * mm - 30 * mm) / col_count

        tabla = Table(table_data, colWidths=[col_width] * col_count)
        tabla.setStyle(
            TableStyle(
                [
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(tabla)

    doc.build(elements)
    return ruta


# ---------------------------------------------------------------------------
# Reporte de Ocupación
# ---------------------------------------------------------------------------


def _generar_pdf_con_graficos(secciones: list, nombre_archivo: str, titulo: str) -> str:
    """
    Genera PDF con múltiples secciones y gráficos incrustados.
    secciones = [
        {"titulo": "KPIs", "datos": [[...]], "grafico": bytes_png},
        {"titulo": "Por Tipo", "datos": [[...]], "grafico": bytes_png},
        ...
    ]
    """
    directorio = _crear_directorio_reportes()
    ruta = os.path.join(directorio, nombre_archivo)

    doc = SimpleDocTemplate(
        ruta,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="TitleCenter",
            parent=styles["Heading1"],
            alignment=1,
            fontSize=16,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Heading2"],
            alignment=0,
            fontSize=12,
            spaceAfter=4,
            spaceBefore=10,
            textColor=colors.HexColor("#2C3E50"),
        )
    )

    elements = []
    elements.append(Paragraph(titulo, styles["TitleCenter"]))
    elements.append(Spacer(1, 4 * mm))

    for i, sec in enumerate(secciones):
        if sec.get("titulo"):
            elements.append(Paragraph(sec["titulo"], styles["SectionTitle"]))
            elements.append(Spacer(1, 2 * mm))

        if sec.get("datos"):
            encabezados = sec["datos"][0]
            filas = sec["datos"][1:]
            table_data = [encabezados] + filas
            col_count = len(encabezados)
            col_width = (210 * mm - 30 * mm) / col_count

            tabla = Table(table_data, colWidths=[col_width] * col_count)
            tabla.setStyle(
                TableStyle(
                    [
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            elements.append(tabla)
            elements.append(Spacer(1, 4 * mm))

        if sec.get("grafico"):
            img_buf = io.BytesIO(sec["grafico"])
            img = Image(img_buf, width=160 * mm, height=80 * mm)
            elements.append(img)
            elements.append(Spacer(1, 6 * mm))

        # Salto de página entre secciones (excepto la última)
        if i < len(secciones) - 1:
            from reportlab.platypus import PageBreak
            elements.append(PageBreak())

    doc.build(elements)
    return ruta


def _generar_xlsx_con_graficos(secciones: list, nombre_archivo: str, encabezados: list) -> str:
    """
    Genera XLSX con múltiples hojas y gráficos.
    secciones = [
        {"hoja": "KPIs", "datos": [...], "grafico": {"tipo": "bar", ...}},
        {"hoja": "Por Tipo", "datos": [...], "grafico": {"tipo": "bar", ...}},
        ...
    ]
    """
    directorio = _crear_directorio_reportes()
    ruta = os.path.join(directorio, nombre_archivo)

    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    for sec in secciones:
        ws = wb.create_sheet(title=sec.get("hoja", "Hoja")[:31])

        # Encabezados
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col_idx, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_idx, value=encabezado)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        cell_align = Alignment(horizontal="left", vertical="center")
        for row_idx, fila in enumerate(sec["datos"], 2):
            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = cell_align
                cell.border = border

        # Auto-ajustar ancho
        for col_idx in range(1, len(encabezados) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, len(sec["datos"]) + 2)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 40)

        # Gráfico si se proporciona
        if sec.get("grafico"):
            grafico_info = sec["grafico"]
            if grafico_info["tipo"] == "bar":
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
            elif grafico_info["tipo"] == "line":
                chart = LineChart()
                chart.style = 10
            else:
                chart = BarChart()

            # Datos del gráfico
            data_ref = Reference(
                ws,
                min_col=grafico_info["data_min_col"],
                max_col=grafico_info["data_max_col"],
                min_row=1,
                max_row=len(sec["datos"]),
            )
            cats_ref = Reference(
                ws,
                min_col=grafico_info["cats_col"],
                min_row=2,
                max_row=len(sec["datos"]),
            )
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.title = grafico_info.get("title", "Gráfico")
            chart.x_axis.title = grafico_info.get("x_title", "")
            chart.y_axis.title = grafico_info.get("y_title", "")

            ws.add_chart(chart, grafico_info.get("posicion", "E2"))

    wb.save(ruta)
    return ruta


# ---------------------------------------------------------------------------
# Reporte de Ocupación (mejorado con 4 secciones + gráficos)
# ---------------------------------------------------------------------------


def generar_ocupacion(
    fecha_inicio: str = None, fecha_fin: str = None, formato: str = "xlsx", creado_por: int = None
) -> dict:
    """
    Genera reporte de ocupación de habitaciones con 4 secciones:
    1. KPIs Principales (Ocupación%, ADR, RevPAR, Ingresos, Reservas, Noches)
    2. Por Tipo de Habitación (Tabla + gráfico barras agrupadas)
    3. Tendencia Diaria (últimos 30 días / período) - línea dual Ocupación% + ADR
    4. Top 5 Habitaciones (ranking ingresos + noches)

    Si fecha_inicio/fecha_fin son None, usa mes anterior completo.
    """
    # Fechas por defecto: mes anterior completo
    if fecha_inicio is None or fecha_fin is None:
        fecha_inicio, fecha_fin = _fecha_default_mes_anterior()

    inicio = _parse_fecha(fecha_inicio)
    fin = _parse_fecha(fecha_fin)

    if inicio > fin:
        raise ValueError("La fecha de inicio debe ser anterior a la fecha de fin.")

    # Reservas en el período (check-in dentro del rango)
    reservas_periodo = (
        db.session.execute(
            select(Reserva).filter(
                Reserva.fecha_entrada >= inicio,
                Reserva.fecha_entrada <= fin,
                Reserva.estado.in_([EstadoReserva.ocupada, EstadoReserva.completada]),
            )
        )
        .scalars()
        .all()
    )

    total_habitaciones = db.session.execute(
        select(func.count()).select_from(Habitacion).filter_by(activo=True)
    ).scalar()
    total_reservas = len(reservas_periodo)
    dias_periodo = (fin - inicio).days + 1

    # --- 1. KPIs PRINCIPALES ---
    noches_totales = sum(r.noches for r in reservas_periodo)
    ingresos_totales = sum(float(r.total) for r in reservas_periodo)
    adr = round(ingresos_totales / noches_totales, 2) if noches_totales > 0 else 0
    revpar = round(ingresos_totales / total_habitaciones, 2) if total_habitaciones > 0 else 0
    ocupacion_general = (
        round(total_reservas / total_habitaciones * 100, 2)
        if total_habitaciones > 0 else 0
    )
    dias_promedio = (
        round(sum(r.noches for r in reservas_periodo) / total_reservas, 2)
        if total_reservas > 0 else 0
    )

    # --- 2. POR TIPO DE HABITACIÓN ---
    reservas_por_tipo = {}
    for tipo in TipoHabitacion:
        reservas_tipo = [r for r in reservas_periodo if r.habitacion.tipo == tipo]
        total_tipo = db.session.execute(
            select(func.count()).select_from(Habitacion).filter_by(tipo=tipo, activo=True)
        ).scalar()
        ingresos_tipo = sum(float(r.total) for r in reservas_tipo)
        adr_tipo = (
            round(ingresos_tipo / sum(r.noches for r in reservas_tipo), 2)
            if reservas_tipo else 0
        )
        revpar_tipo = round(ingresos_tipo / total_tipo, 2) if total_tipo > 0 else 0
        ocupacion_tipo = round(len(reservas_tipo) / total_tipo * 100, 2) if total_tipo > 0 else 0

        reservas_por_tipo[tipo.value] = {
            "habitaciones": total_tipo,
            "reservas": len(reservas_tipo),
            "noches": sum(r.noches for r in reservas_tipo),
            "ingresos": round(ingresos_tipo, 2),
            "ocupacion_pct": ocupacion_tipo,
            "adr": adr_tipo,
            "revpar": revpar_tipo,
        }

    # --- 3. TENDENCIA DIARIA (días en el período) ---
    tendencia_fechas = []
    tendencia_ocupacion = []
    tendencia_adr = []
    tendencia_ingresos = []

    # Pre-calcular reservas por día de check-in
    reservas_por_dia = {}
    for r in reservas_periodo:
        key = r.fecha_entrada
        if key not in reservas_por_dia:
            reservas_por_dia[key] = {"reservas": 0, "noches": 0, "ingresos": 0}
        reservas_por_dia[key]["reservas"] += 1
        reservas_por_dia[key]["noches"] += r.noches
        reservas_por_dia[key]["ingresos"] += float(r.total)

    # Generar serie diaria completa
    dia_actual = inicio
    while dia_actual <= fin:
        tendencia_fechas.append(dia_actual)
        info = reservas_por_dia.get(dia_actual, {"reservas": 0, "noches": 0, "ingresos": 0})
        ocup_dia = (
            round(info["reservas"] / total_habitaciones * 100, 2)
            if total_habitaciones > 0 else 0
        )
        adr_dia = round(info["ingresos"] / info["noches"], 2) if info["noches"] > 0 else 0
        tendencia_ocupacion.append(ocup_dia)
        tendencia_adr.append(adr_dia)
        tendencia_ingresos.append(info["ingresos"])
        dia_actual += timedelta(days=1)

    # --- 4. TOP 5 HABITACIONES ---
    top_ingresos = (
        db.session.query(
            Habitacion.numero,
            func.sum(Reserva.total).label("ingresos"),
            func.sum(Reserva.noches).label("noches"),
        )
        .join(Reserva, Reserva.id_habitacion == Habitacion.id)
        .filter(
            Reserva.fecha_entrada >= inicio,
            Reserva.fecha_entrada <= fin,
            Reserva.estado.in_([EstadoReserva.ocupada, EstadoReserva.completada]),
        )
        .group_by(Habitacion.id, Habitacion.numero)
        .order_by(func.sum(Reserva.total).desc())
        .limit(5)
        .all()
    )

    top_hab_labels = [f"Hab {h.numero}" for h in top_ingresos]
    top_hab_ingresos = [float(h.ingresos) for h in top_ingresos]
    top_hab_noches = [int(h.noches) for h in top_ingresos]

    # --- DETALLE POR HABITACIÓN (para dashboard) ---
    detalle_ocupacion = []
    habitaciones_activas = (
        db.session.execute(select(Habitacion).filter_by(activo=True))
        .scalars().all()
    )
    for h in habitaciones_activas:
        reservas_hab = [r for r in reservas_periodo if r.id_habitacion == h.id]
        noches_ocupadas = sum(r.noches for r in reservas_hab)
        ingreso_hab = sum(float(r.total) for r in reservas_hab)
        total_noches_periodo = dias_periodo
        pct = (
            round(noches_ocupadas / total_noches_periodo * 100, 1)
            if total_noches_periodo > 0 else 0
        )
        detalle_ocupacion.append({
            "numero": h.numero,
            "tipo": h.tipo.value,
            "noches_ocupadas": noches_ocupadas,
            "ingresos": ingreso_hab,
            "ocupacion_pct": min(pct, 100),
        })

    # --- GENERAR GRÁFICOS ---
    # Gráfico 1: Ocupación e Ingresos por Tipo
    labels_tipo = list(reservas_por_tipo.keys())
    ocup_tipo = [reservas_por_tipo[t]["ocupacion_pct"] for t in labels_tipo]
    ing_tipo = [reservas_por_tipo[t]["ingresos"] for t in labels_tipo]
    grafico_tipo = _grafico_ocupacion_por_tipo(labels_tipo, ocup_tipo, ing_tipo)

    # Gráfico 2: Tendencia Diaria
    grafico_tendencia = _grafico_tendencia_diaria(
        tendencia_fechas, tendencia_ocupacion, tendencia_adr,
    )

    # Gráfico 3: Top 5 Habitaciones
    grafico_top = _grafico_top_habitaciones(top_hab_labels, top_hab_ingresos, top_hab_noches)

    # --- PREPARAR SECCIONES PARA PDF/EXCEL ---
    secciones_pdf = []

    # Sección 1: KPIs
    datos_kpis = [
        ["KPI", "Valor"],
        ["Ocupación General (%)", f"{ocupacion_general:.2f}%"],
        ["ADR (COP)", f"${adr:,.2f}"],
        ["RevPAR (COP)", f"${revpar:,.2f}"],
        ["Ingresos Totales (COP)", f"${ingresos_totales:,.2f}"],
        ["Total Reservas", str(total_reservas)],
        ["Noches Totales", str(sum(r.noches for r in reservas_periodo))],
        ["Días Promedio Estancia", f"{dias_promedio:.2f}"],
        ["Período", f"{fecha_inicio} a {fecha_fin} ({dias_periodo} días)"],
    ]
    secciones_pdf.append({"titulo": "1. KPIs Principales", "datos": datos_kpis, "grafico": None})

    # Sección 2: Por Tipo
    datos_tipo = [[
        "Tipo", "Habitaciones", "Reservas", "Noches",
        "Ingresos (COP)", "Ocupación (%)", "ADR (COP)", "RevPAR (COP)",
    ]]
    for tipo, info in reservas_por_tipo.items():
        datos_tipo.append([
            tipo,
            info["habitaciones"],
            info["reservas"],
            info["noches"],
            f"${info['ingresos']:,.2f}",
            f"{info['ocupacion_pct']:.2f}%",
            f"${info['adr']:,.2f}",
            f"${info['revpar']:,.2f}",
        ])
    secciones_pdf.append({
        "titulo": "2. Desglose por Tipo de Habitación",
        "datos": datos_tipo, "grafico": grafico_tipo,
    })

    # Sección 3: Tendencia Diaria
    datos_tendencia = [["Fecha", "Ocupación (%)", "ADR (COP)", "Ingresos (COP)"]]
    for i, f in enumerate(tendencia_fechas):
        datos_tendencia.append([
            f.strftime("%Y-%m-%d"),
            f"{tendencia_ocupacion[i]:.2f}%",
            f"${tendencia_adr[i]:,.2f}",
            f"${tendencia_ingresos[i]:,.2f}",
        ])
    secciones_pdf.append({
        "titulo": "3. Tendencia Diaria (Ocupación + ADR)",
        "datos": datos_tendencia, "grafico": grafico_tendencia,
    })

    # Sección 4: Top 5
    datos_top = [["Habitación", "Ingresos (COP)", "Noches"]]
    for i, h in enumerate(top_ingresos):
        datos_top.append([f"Hab {h.numero}", f"${float(h.ingresos):,.2f}", int(h.noches)])
    secciones_pdf.append({
        "titulo": "4. Top 5 Habitaciones",
        "datos": datos_top, "grafico": grafico_top,
    })

    # --- EXCEL: 4 hojas + gráficos ---
    secciones_xlsx = [
        {
            "hoja": "1. KPIs",
            "datos": datos_kpis,
            "grafico": None,
        },
        {
            "hoja": "2. Por Tipo",
            "datos": datos_tipo,
            "grafico": {
                "tipo": "bar",
                "data_min_col": 5,  # Ingresos
                "data_max_col": 8,  # RevPAR
                "cats_col": 1,      # Tipo
                "title": "Ocupación e Ingresos por Tipo",
                "x_title": "Tipo",
                "y_title": "Valor",
                "posicion": "J2",
            },
        },
        {
            "hoja": "3. Tendencia",
            "datos": datos_tendencia,
            "grafico": {
                "tipo": "line",
                "data_min_col": 2,  # Ocupación
                "data_max_col": 4,  # Ingresos
                "cats_col": 1,      # Fecha
                "title": "Tendencia Diaria: Ocupación y ADR",
                "x_title": "Fecha",
                "y_title": "Valor",
                "posicion": "F2",
            },
        },
        {
            "hoja": "4. Top 5",
            "datos": datos_top,
            "grafico": {
                "tipo": "bar",
                "data_min_col": 2,  # Ingresos
                "data_max_col": 3,  # Noches
                "cats_col": 1,      # Habitación
                "title": "Top 5 Habitaciones",
                "x_title": "Habitación",
                "y_title": "Valor",
                "posicion": "E2",
            },
        },
    ]

    nombre_archivo = f"ocupacion_{fecha_inicio}_{fecha_fin}.{formato}"
    if formato == "pdf":
        ruta = _generar_pdf_con_graficos(secciones_pdf, nombre_archivo, "Reporte de Ocupación")
    elif formato == "json":
        ruta = None
    else:
        ruta = _generar_xlsx_con_graficos(secciones_xlsx, nombre_archivo, [])

    reporte_registrado = None
    if creado_por:
        reporte_registrado = registrar_generado(
            tipo="ocupacion",
            formato=formato,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            archivo_path=ruta or "",
            archivo_nombre=nombre_archivo,
            creado_por=creado_por,
            resumen={
                "total_habitaciones": total_habitaciones,
                "total_reservas": total_reservas,
                "ocupacion_general": ocupacion_general,
                "adr": adr,
                "revpar": revpar,
                "ingresos_totales": ingresos_totales,
                "dias_promedio": dias_promedio,
            },
        )

    return {
        "tipo": "ocupacion",
        "archivo": ruta,
        "formato": formato,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "resumen": {
            "total_habitaciones": total_habitaciones,
            "total_reservas": total_reservas,
            "ocupacion_general": ocupacion_general,
            "adr": adr,
            "revpar": revpar,
            "ingresos_totales": ingresos_totales,
            "dias_promedio": dias_promedio,
            "porcentaje_ocupacion": ocupacion_general,
            "habitaciones_disponibles": total_habitaciones - total_reservas,
        },
        "detalle": detalle_ocupacion,
        "por_tipo": reservas_por_tipo,
        "tendencia_diaria": [
            {
                "fecha": f.strftime("%Y-%m-%d"),
                "ocupacion_pct": tendencia_ocupacion[i],
                "adr": tendencia_adr[i],
                "ingresos": tendencia_ingresos[i],
            }
            for i, f in enumerate(tendencia_fechas)
        ],
        "top_habitaciones": [
            {"numero": f"Hab {h.numero}", "ingresos": float(h.ingresos), "noches": int(h.noches)}
            for h in top_ingresos
        ],
        "historial": reporte_registrado,
    }


# ---------------------------------------------------------------------------
# Reporte de Ingresos
# ---------------------------------------------------------------------------


def generar_ingresos(
    fecha_inicio: str, fecha_fin: str, formato: str = "xlsx", creado_por: int = None
) -> dict:
    """
    Genera reporte de ingresos por período.
    Incluye: total ingresos, desglose subtotal/IVA/servicios,
    ingresos por tipo de habitación.
    """
    inicio = _parse_fecha(fecha_inicio)
    fin = _parse_fecha(fecha_fin)

    if inicio > fin:
        raise ValueError("La fecha de inicio debe ser anterior a la fecha de fin.")

    pagos_aprobados = (
        db.session.execute(
            select(Pago).filter(
                Pago.fecha >= datetime.combine(inicio, datetime.min.time()),
                Pago.fecha <= datetime.combine(fin, datetime.max.time()),
                Pago.estado == EstadoPago.aprobado,
            )
        )
        .scalars()
        .all()
    )

    total_ingresos = sum(_formatear_numero(p.monto) for p in pagos_aprobados)

    ingresos_por_tipo = {}
    subtotal_total = Decimal("0.00")
    iva_total = Decimal("0.00")
    for tipo in TipoHabitacion:
        reservas_tipo = (
            db.session.execute(
                select(Reserva).filter(
                    Reserva.fecha_entrada >= inicio,
                    Reserva.fecha_entrada <= fin,
                    Reserva.habitacion.has(tipo=tipo),
                    Reserva.estado.in_(
                        [
                            EstadoReserva.ocupada,
                            EstadoReserva.completada,
                            EstadoReserva.confirmada,
                        ]
                    ),
                )
            )
            .scalars()
            .all()
        )

        total_tipo = Decimal("0.00")
        servicios_tipo = Decimal("0.00")
        for r in reservas_tipo:
            pagos_reserva = [p for p in pagos_aprobados if p.id_reserva == r.id]
            total_tipo += sum(Decimal(str(p.monto)) for p in pagos_reserva)
            servicios_tipo += sum(
                Decimal(str(s.costo)) for s in r.servicios_adicionales
            )
            # Sumar usando Decimal para evitar mezclar floats y Decimals
            subtotal_total += Decimal(str(r.subtotal))
            iva_total += Decimal(str(r.impuestos))

        ingresos_por_tipo[tipo.value] = {
            "ingresos": _formatear_numero(total_tipo),
            "servicios": _formatear_numero(servicios_tipo),
            "reservas": len(reservas_tipo),
        }

    # Preparar detalle por tipo (para dashboard)
    detalle_ingresos = []
    for tipo, info in ingresos_por_tipo.items():
        detalle_ingresos.append({
            "tipo": tipo,
            "reservas": info["reservas"],
            "servicios": info["servicios"],
            "ingresos": info["ingresos"],
        })

    datos_xlsx = [
        ["Tipo de Habitación", "Reservas", "Servicios Adicionales ($)", "Ingresos ($)"],
    ]
    for tipo, info in ingresos_por_tipo.items():
        datos_xlsx.append(
            [
                tipo,
                info["reservas"],
                info["servicios"],
                info["ingresos"],
            ]
        )
    datos_xlsx.append(["", "", "", ""])
    datos_xlsx.append(["RESUMEN", "", "", ""])
    datos_xlsx.append(["Total ingresos", "", "", round(total_ingresos, 2)])
    datos_xlsx.append(["Subtotal acumulado", "", "", round(subtotal_total, 2)])
    datos_xlsx.append(["IVA acumulado", "", "", round(iva_total, 2)])
    datos_xlsx.append(["Total transacciones", len(pagos_aprobados), "", ""])

    nombre_archivo = f"ingresos_{fecha_inicio}_{fecha_fin}.{formato}"
    if formato == "pdf":
        datos_pdf = [
            ["Tipo", "Reservas", "Servicios ($)", "Ingresos ($)"],
        ]
        for tipo, info in ingresos_por_tipo.items():
            datos_pdf.append(
                [
                    tipo,
                    info["reservas"],
                    info["servicios"],
                    info["ingresos"],
                ]
            )
        datos_pdf.append(["", "", "", ""])
        datos_pdf.append(["RESUMEN GENERAL", "", "", ""])
        datos_pdf.append(["Total ingresos período", "", "", round(total_ingresos, 2)])
        datos_pdf.append(["Subtotal acumulado", "", "", round(subtotal_total, 2)])
        datos_pdf.append(["IVA acumulado", "", "", round(iva_total, 2)])
        datos_pdf.append(["Total transacciones", len(pagos_aprobados), "", ""])
        datos_pdf.append(["Período", f"{fecha_inicio} a {fecha_fin}", "", ""])
        ruta = _generar_pdf(datos_pdf, nombre_archivo, "Reporte de Ingresos")
    elif formato == "json":
        ruta = None
    else:
        ruta = _generar_xlsx(
            datos_xlsx,
            nombre_archivo,
            [
                "Tipo de Habitación",
                "Reservas",
                "Servicios Adicionales ($)",
                "Ingresos ($)",
            ],
        )

    reporte_registrado = None
    if creado_por:
        reporte_registrado = registrar_generado(
            tipo="ingresos",
            formato=formato,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            archivo_path=ruta or "",
            archivo_nombre=nombre_archivo,
            creado_por=creado_por,
            resumen={
                "total_ingresos": round(total_ingresos, 2),
                "subtotal": float(round(subtotal_total, 2)),
                "iva": float(round(iva_total, 2)),
                "transacciones": len(pagos_aprobados),
            },
        )

    return {
        "tipo": "ingresos",
        "archivo": ruta,
        "formato": formato,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "resumen": {
            "total_ingresos": round(total_ingresos, 2),
            "subtotal": float(round(subtotal_total, 2)),
            "iva": float(round(iva_total, 2)),
            "transacciones": len(pagos_aprobados),
        },
        "detalle": detalle_ingresos,
        "historial": reporte_registrado,
    }


# ---------------------------------------------------------------------------
# Reporte de Estadísticas
# ---------------------------------------------------------------------------


def generar_estadisticas(
    fecha_inicio: str, fecha_fin: str, formato: str = "xlsx", creado_por: int = None
) -> dict:
    """
    Genera reporte de estadísticas generales.
    Incluye: reservas por estado, top 5 habitaciones más reservadas,
    huéspedes con más reservas.
    """
    inicio = _parse_fecha(fecha_inicio)
    fin = _parse_fecha(fecha_fin)

    if inicio > fin:
        raise ValueError("La fecha de inicio debe ser anterior a la fecha de fin.")

    reservas_periodo = (
        db.session.execute(
            select(Reserva).filter(
                Reserva.fecha_entrada >= inicio,
                Reserva.fecha_entrada <= fin,
            )
        )
        .scalars()
        .all()
    )

    reservas_por_estado = {}
    for estado in EstadoReserva:
        count = sum(1 for r in reservas_periodo if r.estado == estado)
        reservas_por_estado[estado.value] = count

    top_habitaciones = (
        db.session.query(
            Habitacion.numero,
            Habitacion.tipo,
            func.count(Reserva.id).label("total"),
        )
        .join(Reserva, Reserva.id_habitacion == Habitacion.id)
        .filter(
            Reserva.fecha_entrada >= inicio,
            Reserva.fecha_entrada <= fin,
        )
        .group_by(Habitacion.id, Habitacion.numero, Habitacion.tipo)
        .order_by(func.count(Reserva.id).desc())
        .limit(5)
        .all()
    )

    top_huespedes = (
        db.session.query(
            Huesped.documento_id,
            Huesped.tipo_documento,
            Huesped.id.label("huesped_id"),
            func.count(Reserva.id).label("total"),
        )
        .join(Reserva, Reserva.id_huesped == Huesped.id)
        .filter(
            Reserva.fecha_entrada >= inicio,
            Reserva.fecha_entrada <= fin,
        )
        .group_by(Huesped.id, Huesped.documento_id, Huesped.tipo_documento)
        .order_by(func.count(Reserva.id).desc())
        .limit(5)
        .all()
    )

    datos_xlsx = [
        ["RESERVAS POR ESTADO", "", "", ""],
        ["Estado", "Cantidad", "", ""],
    ]
    for estado, cantidad in reservas_por_estado.items():
        datos_xlsx.append([estado, cantidad, "", ""])
    datos_xlsx.append(["", "", "", ""])

    datos_xlsx.append(["TOP 5 HABITACIONES MÁS RESERVADAS", "", "", ""])
    datos_xlsx.append(["Habitación", "Tipo", "Reservas", ""])
    for h in top_habitaciones:
        datos_xlsx.append([h.numero, h.tipo.value, h.total, ""])
    datos_xlsx.append(["", "", "", ""])

    datos_xlsx.append(["TOP 5 HUÉSPEDES CON MÁS RESERVAS", "", "", ""])
    datos_xlsx.append(["Documento", "Tipo", "Nombre", "Reservas"])
    for h in top_huespedes:
        huesped_obj = db.session.get(Huesped, h.huesped_id)
        nombre = (
            f"{huesped_obj.usuario.nombre} {huesped_obj.usuario.apellido}"
            if huesped_obj and huesped_obj.usuario
            else "N/A"
        )
        datos_xlsx.append([h.documento_id, h.tipo_documento, nombre, h.total])

    nombre_archivo = f"estadisticas_{fecha_inicio}_{fecha_fin}.{formato}"
    if formato == "pdf":
        datos_pdf = [["Sección", "Detalle", "Valor", ""]]
        datos_pdf.append(["Total reservas", str(len(reservas_periodo)), "", ""])
        datos_pdf.append(["", "", "", ""])
        datos_pdf.append(["ESTADO", "CANTIDAD", "", ""])
        for estado, cantidad in reservas_por_estado.items():
            datos_pdf.append([estado, str(cantidad), "", ""])
        datos_pdf.append(["", "", "", ""])
        datos_pdf.append(["HABITACIÓN", "TIPO", "RESERVAS", ""])
        for h in top_habitaciones:
            datos_pdf.append([h.numero, h.tipo.value, str(h.total), ""])
        datos_pdf.append(["", "", "", ""])
        datos_pdf.append(["DOCUMENTO", "NOMBRE", "RESERVAS", ""])
        for h in top_huespedes:
            huesped_obj = db.session.get(Huesped, h.huesped_id)
            nombre = (
                f"{huesped_obj.usuario.nombre} {huesped_obj.usuario.apellido}"
                if huesped_obj and huesped_obj.usuario
                else "N/A"
            )
            datos_pdf.append([h.documento_id, nombre, str(h.total), ""])
        ruta = _generar_pdf(datos_pdf, nombre_archivo, "Reporte de Estadísticas")
    else:
        ruta = _generar_xlsx(
            datos_xlsx,
            nombre_archivo,
            ["Concepto", "Detalle 1", "Detalle 2", "Valor"],
        )

    reporte_registrado = None
    if creado_por:
        reporte_registrado = registrar_generado(
            tipo="estadisticas",
            formato=formato,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            archivo_path=ruta,
            archivo_nombre=nombre_archivo,
            creado_por=creado_por,
            resumen={
                "total_reservas": len(reservas_periodo),
                "reservas_por_estado": reservas_por_estado,
            },
        )

    return {
        "tipo": "estadisticas",
        "archivo": ruta,
        "formato": formato,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "resumen": {
            "total_reservas": len(reservas_periodo),
            "reservas_por_estado": reservas_por_estado,
        },
        "historial": reporte_registrado,
    }


def registrar_generado(
    tipo: str,
    formato: str,
    fecha_inicio: str,
    fecha_fin: str,
    archivo_path: str,
    archivo_nombre: str,
    creado_por: int,
    resumen: dict = None,
) -> dict:
    """Registra un reporte generado en el historial."""
    from app.models.reporte import ReporteGenerado
    from datetime import date

    reporte = ReporteGenerado(
        tipo=tipo,
        formato=formato,
        fecha_inicio=date.fromisoformat(fecha_inicio),
        fecha_fin=date.fromisoformat(fecha_fin),
        archivo_path=archivo_path,
        archivo_nombre=archivo_nombre,
        creado_por=creado_por,
        resumen=resumen,
    )
    db.session.add(reporte)
    db.session.flush()
    return reporte.to_dict()


def listar_historial(filtros=None):
    """Lista el historial de reportes generados con filtros opcionales."""
    from app.models.reporte import ReporteGenerado

    query = ReporteGenerado.query

    if filtros:
        if filtros.get("tipo"):
            query = query.filter_by(tipo=filtros["tipo"])

        if filtros.get("formato"):
            query = query.filter_by(formato=filtros["formato"])

        if filtros.get("creado_por"):
            query = query.filter_by(creado_por=int(filtros["creado_por"]))

        if filtros.get("fecha_desde"):
            from datetime import date

            fecha = date.fromisoformat(filtros["fecha_desde"])
            query = query.filter(ReporteGenerado.created_at >= fecha)

        if filtros.get("fecha_hasta"):
            from datetime import date

            fecha = date.fromisoformat(filtros["fecha_hasta"])
            query = query.filter(ReporteGenerado.created_at <= fecha)

    reportes = query.order_by(ReporteGenerado.created_at.desc()).all()
    return [r.to_dict() for r in reportes]
